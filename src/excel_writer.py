"""
excel_writer.py
Writes the list of analyzed emails into a clean, formatted Excel file.
Emails are already sorted by date when they arrive here (sorted in eml_reader.py).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


HEADERS = ["Date", "Sent From", "Sent To", "Subject", "Keywords", "Summary"]

COLUMN_WIDTHS = {
    "A": 18,  # Date
    "B": 28,  # Sent From
    "C": 28,  # Sent To
    "D": 35,  # Subject
    "E": 40,  # Keywords
    "F": 60,  # Summary
}

HEADER_FILL_COLOR = "1F4E79"
HEADER_FONT_COLOR = "FFFFFF"
ROW_ALT_COLOR = "D6E4F0"


def write_excel(emails: list[dict], output_path: str):
    """Write analyzed email data to a formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Email Recap"

    _write_headers(ws)
    _write_rows(ws, emails)
    _apply_column_widths(ws)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel disimpan ke: {output_path}")


def _write_headers(ws):
    ws.append(HEADERS)
    for col_num in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR, name="Arial", size=11)
        cell.fill = PatternFill("solid", start_color=HEADER_FILL_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 20


def _write_rows(ws, emails: list[dict]):
    for i, email_data in enumerate(emails, start=2):
        # Use clean parsed date if available, fallback to raw date string
        display_date = email_data.get("date_parsed") or email_data.get("date", "")

        row = [
            display_date,
            email_data.get("sent_from", ""),
            email_data.get("sent_to", ""),
            email_data.get("subject", ""),
            email_data.get("keywords", ""),
            email_data.get("summary", ""),
        ]
        ws.append(row)

        if i % 2 == 0:
            for col_num in range(1, len(HEADERS) + 1):
                ws.cell(row=i, column=col_num).fill = PatternFill("solid", start_color=ROW_ALT_COLOR)

        for col_num in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=i, column=col_num)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.row_dimensions[i].height = 50


def _apply_column_widths(ws):
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
